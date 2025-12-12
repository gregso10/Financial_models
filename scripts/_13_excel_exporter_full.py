# In file: scripts/_12_excel_exporter_full.py

"""
Full Financial Model Excel Exporter
Generates a complete Excel model with:
- Assumptions sheet with named ranges
- Monthly P&L, BS, CF, Loan Schedule with formulas
- Yearly summary sheets aggregating monthly data
- All cells linked via formulas (change assumptions → model recalculates)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.workbook.defined_name import DefinedName
from typing import Dict, List, Tuple, Optional
from io import BytesIO


class ExcelExporterFull:
    """
    Exports complete financial model to Excel with full formula linkage.
    Monthly granularity with yearly aggregations.
    """
    
    # ===== STYLES =====
    BLUE_INPUT = Font(color="0000FF", bold=False)
    BLACK_CALC = Font(color="000000", bold=False)
    BOLD = Font(bold=True)
    BOLD_TOTAL = Font(bold=True, color="000000")
    
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
    SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
    TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
    INPUT_FILL = PatternFill("solid", fgColor="FFFFCC")  # Light yellow for inputs
    
    THIN_BORDER = Border(bottom=Side(style='thin', color='000000'))
    THICK_BORDER = Border(top=Side(style='medium', color='000000'),
                          bottom=Side(style='double', color='000000'))
    
    # Number formats
    NUM_FORMAT = '#,##0;(#,##0);"-"'
    NUM_FORMAT_2 = '#,##0.00;(#,##0.00);"-"'
    PCT_FORMAT = '0.00%;(0.00%);"-"'
    PCT_FORMAT_1 = '0.0%;(0.0%);"-"'
    
    def __init__(self, params, pnl_df: pd.DataFrame, bs_df: pd.DataFrame,
                 cf_df: pd.DataFrame, loan_schedule: pd.DataFrame,
                 investment_metrics: Dict):
        self.params = params
        self.pnl_df = pnl_df
        self.bs_df = bs_df
        self.cf_df = cf_df
        self.loan_schedule = loan_schedule
        self.metrics = investment_metrics or {}
        self.wb = Workbook()
        
        # Track named ranges for cross-references
        self.named_ranges: Dict[str, str] = {}
        
        # Track P&L column positions (populated during monthly P&L creation)
        self.pnl_col_map: Dict[str, int] = {}
        
        # Determine lease type from params
        self.lease_type = getattr(params, 'lease_type_used', 'furnished_1yr')
        
    def export(self) -> BytesIO:
        """Generate complete Excel model."""
        self.wb.remove(self.wb.active)
        
        # Create sheets in order
        self._create_assumptions_sheet()
        self._create_monthly_loan_sheet()
        self._create_monthly_pnl_sheet()
        self._create_monthly_cf_sheet()
        self._create_monthly_bs_sheet()
        self._create_yearly_pnl_sheet()
        self._create_yearly_cf_sheet()
        self._create_yearly_bs_sheet()
        self._create_metrics_sheet()
        
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    def _add_named_range(self, name: str, sheet: str, cell: str):
        """Add a named range to workbook."""
        ref = f"'{sheet}'!{cell}"
        try:
            defn = DefinedName(name, attr_text=ref)
            self.wb.defined_names[name] = defn
            self.named_ranges[name] = ref
        except Exception:
            pass  # Skip if name already exists
    
    def _create_assumptions_sheet(self):
        """Create Assumptions sheet with named ranges for all inputs."""
        ws = self.wb.create_sheet("Assumptions")
        
        # Get rental assumptions for selected lease type
        rental = self.params.rental_assumptions.get(self.lease_type, {})
        
        # Structure: (label, value, named_range_name, format_type)
        # format_type: 'num', 'pct', 'int', 'text'
        assumptions = [
            # === PROPERTY ===
            ("PROPERTY & ACQUISITION", None, None, None),
            ("Property Price (FAI)", self.params.property_price, "PropertyPrice", "num"),
            ("Property Size (sqm)", self.params.property_size_sqm, "PropertySize", "num"),
            ("Agency Fees %", self.params.agency_fees_percentage, "AgencyFeesPct", "pct"),
            ("Notary Fees %", self.params.notary_fees_percentage_estimate, "NotaryFeesPct", "pct"),
            ("Initial Renovation", self.params.initial_renovation_costs, "RenovationCost", "num"),
            ("Furnishing Costs", self.params.furnishing_costs, "FurnishingCost", "num"),
            ("", None, None, None),
            
            # === FINANCING ===
            ("FINANCING", None, None, None),
            ("Loan Percentage", self.params.loan_percentage, "LoanPct", "pct"),
            ("Loan Interest Rate (Annual)", self.params.loan_interest_rate, "LoanRate", "pct"),
            ("Loan Duration (Years)", self.params.loan_duration_years, "LoanYears", "int"),
            ("Loan Insurance Rate", self.params.loan_insurance_rate, "LoanInsuranceRate", "pct"),
            ("", None, None, None),
            
            # === RENTAL ===
            ("RENTAL ASSUMPTIONS", None, None, None),
            ("Lease Type", self.lease_type, "LeaseType", "text"),
        ]
        
        # Add lease-specific assumptions
        if self.lease_type == "airbnb":
            assumptions.extend([
                ("Daily Rate", rental.get("daily_rate", 0), "DailyRate", "num"),
                ("Occupancy Rate", rental.get("occupancy_rate", 0.7), "OccupancyRate", "pct"),
                ("Rent Growth Rate", rental.get("rent_growth_rate", 0.02), "RentGrowth", "pct"),
            ])
        else:
            assumptions.extend([
                ("Monthly Rent per sqm", rental.get("monthly_rent_sqm", 0), "RentPerSqm", "num"),
                ("Vacancy Rate (Annual)", rental.get("vacancy_rate", 0.08), "VacancyRate", "pct"),
                ("Rent Growth Rate", rental.get("rent_growth_rate", 0.015), "RentGrowth", "pct"),
            ])
        
        assumptions.extend([
            ("", None, None, None),
            
            # === OPERATING EXPENSES ===
            ("OPERATING EXPENSES", None, None, None),
            ("Property Tax (Yearly)", self.params.property_tax_yearly, "PropertyTax", "num"),
            ("Condo Fees (Monthly)", self.params.condo_fees_monthly, "CondoFees", "num"),
            ("PNO Insurance (Yearly)", self.params.pno_insurance_yearly, "PNOInsurance", "num"),
            ("Maintenance % of Rent", self.params.maintenance_percentage_rent, "MaintenancePct", "pct"),
            ("Management Fee %", self.params.management_fees_percentage_rent.get(self.lease_type, 0.07), "ManagementPct", "pct"),
            ("Expenses Growth Rate", self.params.expenses_growth_rate, "ExpensesGrowth", "pct"),
            ("", None, None, None),
            
            # === FISCAL ===
            ("FISCAL PARAMETERS", None, None, None),
            ("Fiscal Regime", self.params.fiscal_regime, "FiscalRegime", "text"),
            ("Income Tax Bracket (TMI)", self.params.personal_income_tax_bracket, "TMI", "pct"),
            ("Social Contributions Rate", self.params.social_contributions_rate, "SocialRate", "pct"),
            ("Property Amortization Years", self.params.lmnp_amortization_property_years, "AmortPropYears", "int"),
            ("Furnishing Amortization Years", self.params.lmnp_amortization_furnishing_years, "AmortFurnYears", "int"),
            ("Renovation Amortization Years", self.params.lmnp_amortization_renovation_years, "AmortRenoYears", "int"),
            ("", None, None, None),
            
            # === EXIT ===
            ("EXIT PARAMETERS", None, None, None),
            ("Holding Period (Years)", self.params.holding_period_years, "HoldingYears", "int"),
            ("Property Growth Rate", self.params.property_value_growth_rate, "PropertyGrowth", "pct"),
            ("Selling Fees %", self.params.exit_selling_fees_percentage, "SellingFeesPct", "pct"),
            ("", None, None, None),
            
            # === CALCULATED VALUES (for reference) ===
            ("CALCULATED VALUES", None, None, None),
            ("Total Acquisition Cost", "=PropertyPrice*(1+NotaryFeesPct)+RenovationCost+FurnishingCost", "TotalAcqCost", "formula"),
            ("Loan Amount", "=TotalAcqCost*LoanPct", "LoanAmount", "formula"),
            ("Initial Equity", "=TotalAcqCost-LoanAmount", "InitialEquity", "formula"),
            ("Monthly Loan Rate", "=LoanRate/12", "MonthlyRate", "formula"),
            ("Number of Payments", "=LoanYears*12", "NumPayments", "formula"),
            ("Monthly Payment", "=IF(LoanAmount>0,-PMT(MonthlyRate,NumPayments,LoanAmount),0)", "MonthlyPayment", "formula"),
            ("Land Value (15%)", "=PropertyPrice/(1+AgencyFeesPct)*0.15", "LandValue", "formula"),
            ("Amortizable Property Value", "=PropertyPrice/(1+AgencyFeesPct)*0.85", "AmortPropValue", "formula"),
            ("Monthly Property Amort", "=AmortPropValue/AmortPropYears/12", "MonthlyPropAmort", "formula"),
            ("Monthly Furnishing Amort", "=IF(FurnishingCost>0,FurnishingCost/AmortFurnYears/12,0)", "MonthlyFurnAmort", "formula"),
            ("Monthly Renovation Amort", "=IF(RenovationCost>0,RenovationCost/AmortRenoYears/12,0)", "MonthlyRenoAmort", "formula"),
        ])
        
        # Write to sheet
        row = 1
        for label, value, name, fmt in assumptions:
            if value is None and label and label != "":
                # Section header
                cell = ws.cell(row=row, column=1, value=label)
                cell.font = self.HEADER_FONT
                cell.fill = self.HEADER_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            elif label == "":
                pass  # Empty row
            else:
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2)
                
                if fmt == "formula":
                    cell.value = value  # It's already a formula string
                    cell.font = self.BLACK_CALC
                    cell.number_format = self.NUM_FORMAT
                else:
                    cell.value = value
                    cell.font = self.BLUE_INPUT
                    cell.fill = self.INPUT_FILL
                    
                    if fmt == "pct":
                        cell.number_format = self.PCT_FORMAT
                    elif fmt == "num":
                        cell.number_format = self.NUM_FORMAT
                    elif fmt == "int":
                        cell.number_format = "0"
                
                # Create named range
                if name:
                    self._add_named_range(name, "Assumptions", f"$B${row}")
            
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 18
    
    def _create_monthly_loan_sheet(self):
        """Create monthly loan amortization with PMT/IPMT/PPMT formulas."""
        ws = self.wb.create_sheet("Loan_Monthly")
        
        num_months = self.params.holding_period_years * 12
        loan_months = self.params.loan_duration_years * 12
        
        # Headers
        headers = ["Month", "Year", "Beginning Balance", "Payment", "Interest", "Principal", "Ending Balance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows with formulas
        for month in range(1, num_months + 1):
            row = month + 1
            
            # Month number
            ws.cell(row=row, column=1, value=month)
            
            # Year
            ws.cell(row=row, column=2, value=f"=INT((A{row}-1)/12)+1")
            
            # Beginning Balance
            if month == 1:
                ws.cell(row=row, column=3, value="=LoanAmount")
            else:
                ws.cell(row=row, column=3, value=f"=G{row-1}")  # Previous ending balance
            
            # Payment (0 if beyond loan term)
            ws.cell(row=row, column=4, value=f"=IF(A{row}<=NumPayments,MonthlyPayment,0)")
            
            # Interest
            ws.cell(row=row, column=5, value=f"=IF(A{row}<=NumPayments,-IPMT(MonthlyRate,A{row},NumPayments,LoanAmount),0)")
            
            # Principal
            ws.cell(row=row, column=6, value=f"=IF(A{row}<=NumPayments,-PPMT(MonthlyRate,A{row},NumPayments,LoanAmount),0)")
            
            # Ending Balance
            ws.cell(row=row, column=7, value=f"=MAX(0,C{row}-F{row})")
            
            # Format
            for col in range(3, 8):
                ws.cell(row=row, column=col).number_format = self.NUM_FORMAT
        
        # Column widths
        widths = [8, 6, 18, 14, 14, 14, 18]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_monthly_pnl_sheet(self):
        """Create monthly P&L with formulas linked to Assumptions."""
        ws = self.wb.create_sheet("PnL_Monthly")
        
        num_months = self.params.holding_period_years * 12
        
        # Build structure dynamically and track columns
        # Each entry: (label, formula_builder_func, is_subtotal)
        # We'll build formulas after knowing column positions
        
        headers = ["Month", "Year", "Gross Potential Rent"]
        if self.lease_type == "airbnb":
            headers.append("Occupancy Adj")
        else:
            headers.append("Vacancy Loss")
        headers.extend([
            "Gross Operating Income",
            "Property Tax", "Condo Fees", "PNO Insurance", 
            "Maintenance", "Management Fees"
        ])
        if self.lease_type == "airbnb":
            headers.append("Airbnb Costs")
        headers.extend([
            "Total OpEx", "NOI",
            "Loan Interest", "Loan Insurance",
            "Depr Property", "Depr Furnishing", "Depr Renovation", "Total Depreciation",
            "Taxable Income",
            "Income Tax", "Social Contributions", "Total Taxes",
            "Net Income"
        ])
        
        # Write headers and build column map
        col_map = {}
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            col_map[header] = col
        
        # Helper to get column letter
        def c(name):
            return get_column_letter(col_map[name])
        
        # Write monthly data
        for month in range(1, num_months + 1):
            r = month + 1  # Data row
            loan_r = month + 1
            
            # Month & Year
            ws.cell(row=r, column=col_map["Month"], value=month)
            ws.cell(row=r, column=col_map["Year"], value=f"=INT((A{r}-1)/12)+1")
            
            # Revenue
            if self.lease_type == "airbnb":
                ws.cell(row=r, column=col_map["Gross Potential Rent"], 
                       value=f"=DailyRate*30*(1+RentGrowth)^({c('Year')}{r}-1)")
                ws.cell(row=r, column=col_map["Occupancy Adj"],
                       value=f"=-{c('Gross Potential Rent')}{r}*(1-OccupancyRate)")
                ws.cell(row=r, column=col_map["Gross Operating Income"],
                       value=f"={c('Gross Potential Rent')}{r}+{c('Occupancy Adj')}{r}")
            else:
                ws.cell(row=r, column=col_map["Gross Potential Rent"],
                       value=f"=RentPerSqm*PropertySize*(1+RentGrowth)^({c('Year')}{r}-1)")
                ws.cell(row=r, column=col_map["Vacancy Loss"],
                       value=f"=-{c('Gross Potential Rent')}{r}*VacancyRate/12")
                ws.cell(row=r, column=col_map["Gross Operating Income"],
                       value=f"={c('Gross Potential Rent')}{r}+{c('Vacancy Loss')}{r}")
            
            # Operating Expenses (all negative)
            goi = f"{c('Gross Operating Income')}{r}"
            ws.cell(row=r, column=col_map["Property Tax"],
                   value=f"=-PropertyTax/12*(1+ExpensesGrowth)^({c('Year')}{r}-1)")
            ws.cell(row=r, column=col_map["Condo Fees"],
                   value=f"=-CondoFees*(1+ExpensesGrowth)^({c('Year')}{r}-1)")
            ws.cell(row=r, column=col_map["PNO Insurance"],
                   value=f"=-PNOInsurance/12*(1+ExpensesGrowth)^({c('Year')}{r}-1)")
            ws.cell(row=r, column=col_map["Maintenance"],
                   value=f"=-{goi}*MaintenancePct")
            ws.cell(row=r, column=col_map["Management Fees"],
                   value=f"=-{goi}*ManagementPct")
            
            if self.lease_type == "airbnb":
                ws.cell(row=r, column=col_map["Airbnb Costs"],
                       value=f"=-{goi}*0.15")
                opex_range = f"{c('Property Tax')}{r}:{c('Airbnb Costs')}{r}"
            else:
                opex_range = f"{c('Property Tax')}{r}:{c('Management Fees')}{r}"
            
            ws.cell(row=r, column=col_map["Total OpEx"],
                   value=f"=SUM({opex_range})")
            ws.cell(row=r, column=col_map["NOI"],
                   value=f"={goi}+{c('Total OpEx')}{r}")
            
            # Financing & Depreciation
            ws.cell(row=r, column=col_map["Loan Interest"],
                   value=f"=-Loan_Monthly!E{loan_r}")
            ws.cell(row=r, column=col_map["Loan Insurance"],
                   value=f"=-IF(A{r}<=NumPayments,LoanAmount*LoanInsuranceRate/12,0)")
            ws.cell(row=r, column=col_map["Depr Property"],
                   value=f"=-IF({c('Year')}{r}<=AmortPropYears,MonthlyPropAmort,0)")
            ws.cell(row=r, column=col_map["Depr Furnishing"],
                   value=f"=-IF({c('Year')}{r}<=AmortFurnYears,MonthlyFurnAmort,0)")
            ws.cell(row=r, column=col_map["Depr Renovation"],
                   value=f"=-IF({c('Year')}{r}<=AmortRenoYears,MonthlyRenoAmort,0)")
            ws.cell(row=r, column=col_map["Total Depreciation"],
                   value=f"=SUM({c('Depr Property')}{r}:{c('Depr Renovation')}{r})")
            
            # Taxable Income
            ws.cell(row=r, column=col_map["Taxable Income"],
                   value=f"={c('NOI')}{r}+{c('Loan Interest')}{r}+{c('Loan Insurance')}{r}+{c('Total Depreciation')}{r}")
            
            # Taxes
            taxable = f"{c('Taxable Income')}{r}"
            ws.cell(row=r, column=col_map["Income Tax"],
                   value=f"=-MAX(0,{taxable})*TMI")
            ws.cell(row=r, column=col_map["Social Contributions"],
                   value=f"=-MAX(0,{taxable})*SocialRate")
            ws.cell(row=r, column=col_map["Total Taxes"],
                   value=f"={c('Income Tax')}{r}+{c('Social Contributions')}{r}")
            
            # Net Income
            ws.cell(row=r, column=col_map["Net Income"],
                   value=f"={taxable}+{c('Total Taxes')}{r}")
            
            # Format all numeric cells
            for col in range(3, len(headers) + 1):
                ws.cell(row=r, column=col).number_format = self.NUM_FORMAT
        
        # Store column map for other sheets to reference
        self.pnl_col_map = col_map
        
        # Column widths
        ws.column_dimensions['A'].width = 7
        ws.column_dimensions['B'].width = 5
        for col in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 13
    
    def _create_monthly_cf_sheet(self):
        """Create monthly Cash Flow with formulas."""
        ws = self.wb.create_sheet("CF_Monthly")
        
        num_months = self.params.holding_period_years * 12
        
        # Get P&L column letters (set during PnL creation)
        pnl_net_income_col = get_column_letter(self.pnl_col_map.get("Net Income", 23))
        pnl_depr_col = get_column_letter(self.pnl_col_map.get("Total Depreciation", 18))
        
        # Headers
        headers = [
            "Month", "Year",
            "Net Income", "Add: Depreciation", "CFO",
            "Acquisition", "CFI",
            "Loan Proceeds", "Equity Injected", "Principal Repayment", "CFF",
            "Net Change", "Beginning Cash", "Ending Cash"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Monthly data with formulas
        for month in range(1, num_months + 1):
            row = month + 1
            pnl_row = month + 1
            loan_row = month + 1
            
            # Month, Year
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=f"=INT((A{row}-1)/12)+1")
            
            # Net Income (from P&L)
            ws.cell(row=row, column=3, value=f"=PnL_Monthly!{pnl_net_income_col}{pnl_row}")
            
            # Add: Depreciation (flip sign - was negative in P&L)
            ws.cell(row=row, column=4, value=f"=-PnL_Monthly!{pnl_depr_col}{pnl_row}")
            
            # CFO
            ws.cell(row=row, column=5, value=f"=C{row}+D{row}")
            
            # Acquisition (Month 1 only)
            ws.cell(row=row, column=6, value=f"=IF(A{row}=1,-TotalAcqCost,0)")
            
            # CFI
            ws.cell(row=row, column=7, value=f"=F{row}")
            
            # Loan Proceeds (Month 1 only)
            ws.cell(row=row, column=8, value=f"=IF(A{row}=1,LoanAmount,0)")
            
            # Equity Injected (Month 1 only)
            ws.cell(row=row, column=9, value=f"=IF(A{row}=1,InitialEquity,0)")
            
            # Principal Repayment (from Loan sheet)
            ws.cell(row=row, column=10, value=f"=-Loan_Monthly!F{loan_row}")
            
            # CFF
            ws.cell(row=row, column=11, value=f"=H{row}+I{row}+J{row}")
            
            # Net Change
            ws.cell(row=row, column=12, value=f"=E{row}+G{row}+K{row}")
            
            # Beginning Cash
            if month == 1:
                ws.cell(row=row, column=13, value=0)
            else:
                ws.cell(row=row, column=13, value=f"=N{row-1}")
            
            # Ending Cash
            ws.cell(row=row, column=14, value=f"=M{row}+L{row}")
            
            # Format
            for col in range(3, 15):
                ws.cell(row=row, column=col).number_format = self.NUM_FORMAT
        
        # Column widths
        for col in range(1, 15):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_monthly_bs_sheet(self):
        """Create monthly Balance Sheet with formulas."""
        ws = self.wb.create_sheet("BS_Monthly")
        
        num_months = self.params.holding_period_years * 12
        
        # Get P&L column letter for Net Income
        pnl_net_income_col = get_column_letter(self.pnl_col_map.get("Net Income", 23))
        
        headers = [
            "Month", "Year",
            "Property Gross", "Property Accum Depr", "Property Net",
            "Renovation Gross", "Renovation Accum Depr", "Renovation Net",
            "Furnishing Gross", "Furnishing Accum Depr", "Furnishing Net",
            "Total Fixed Assets", "Cash", "Total Assets",
            "Loan Balance", "Total Liabilities",
            "Initial Equity", "Retained Earnings", "Total Equity",
            "Total L&E", "Balance Check"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Row 2 = Month 0 (Initial)
        ws.cell(row=2, column=1, value=0)
        ws.cell(row=2, column=2, value=0)
        ws.cell(row=2, column=3, value="=PropertyPrice*(1+NotaryFeesPct)")  # Property Gross
        ws.cell(row=2, column=4, value=0)  # Accum Depr
        ws.cell(row=2, column=5, value="=C2-D2")  # Net
        ws.cell(row=2, column=6, value="=RenovationCost")  # Renovation Gross
        ws.cell(row=2, column=7, value=0)
        ws.cell(row=2, column=8, value="=F2-G2")
        ws.cell(row=2, column=9, value="=FurnishingCost")  # Furnishing Gross
        ws.cell(row=2, column=10, value=0)
        ws.cell(row=2, column=11, value="=I2-J2")
        ws.cell(row=2, column=12, value="=E2+H2+K2")  # Total Fixed
        ws.cell(row=2, column=13, value=0)  # Cash
        ws.cell(row=2, column=14, value="=L2+M2")  # Total Assets
        ws.cell(row=2, column=15, value="=LoanAmount")  # Loan Balance
        ws.cell(row=2, column=16, value="=O2")  # Total Liabilities
        ws.cell(row=2, column=17, value="=InitialEquity")  # Initial Equity
        ws.cell(row=2, column=18, value=0)  # Retained Earnings
        ws.cell(row=2, column=19, value="=Q2+R2")  # Total Equity
        ws.cell(row=2, column=20, value="=P2+S2")  # Total L&E
        ws.cell(row=2, column=21, value="=N2-T2")  # Balance Check
        
        for col in range(3, 22):
            ws.cell(row=2, column=col).number_format = self.NUM_FORMAT
        
        # Months 1 to num_months
        for month in range(1, num_months + 1):
            row = month + 2
            prev_row = row - 1
            pnl_row = month + 1
            cf_row = month + 1
            loan_row = month + 1
            
            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=f"=INT((A{row}-1)/12)+1")
            
            # Property
            ws.cell(row=row, column=3, value=f"=C{prev_row}")  # Gross unchanged
            ws.cell(row=row, column=4, value=f"=D{prev_row}+IF(B{row}<=AmortPropYears,MonthlyPropAmort,0)")
            ws.cell(row=row, column=5, value=f"=C{row}-D{row}")
            
            # Renovation
            ws.cell(row=row, column=6, value=f"=F{prev_row}")
            ws.cell(row=row, column=7, value=f"=G{prev_row}+IF(B{row}<=AmortRenoYears,MonthlyRenoAmort,0)")
            ws.cell(row=row, column=8, value=f"=F{row}-G{row}")
            
            # Furnishing
            ws.cell(row=row, column=9, value=f"=I{prev_row}")
            ws.cell(row=row, column=10, value=f"=J{prev_row}+IF(B{row}<=AmortFurnYears,MonthlyFurnAmort,0)")
            ws.cell(row=row, column=11, value=f"=I{row}-J{row}")
            
            # Totals
            ws.cell(row=row, column=12, value=f"=E{row}+H{row}+K{row}")
            ws.cell(row=row, column=13, value=f"=CF_Monthly!N{cf_row}")  # Cash from CF
            ws.cell(row=row, column=14, value=f"=L{row}+M{row}")
            ws.cell(row=row, column=15, value=f"=Loan_Monthly!G{loan_row}")  # Loan balance
            ws.cell(row=row, column=16, value=f"=O{row}")
            ws.cell(row=row, column=17, value="=InitialEquity")
            ws.cell(row=row, column=18, value=f"=R{prev_row}+PnL_Monthly!{pnl_net_income_col}{pnl_row}")  # Cumulative Net Income
            ws.cell(row=row, column=19, value=f"=Q{row}+R{row}")
            ws.cell(row=row, column=20, value=f"=P{row}+S{row}")
            ws.cell(row=row, column=21, value=f"=N{row}-T{row}")
            
            for col in range(3, 22):
                ws.cell(row=row, column=col).number_format = self.NUM_FORMAT
        
        # Column widths
        for col in range(1, 22):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_yearly_pnl_sheet(self):
        """Create yearly P&L aggregating monthly data via SUMIF."""
        ws = self.wb.create_sheet("PnL_Yearly")
        
        years = self.params.holding_period_years
        
        # Build row definitions using actual column map from monthly P&L
        # (label, monthly_col_name, is_total)
        pnl_rows = [
            ("Gross Potential Rent", "Gross Potential Rent", False),
        ]
        if self.lease_type == "airbnb":
            pnl_rows.append(("Occupancy Adjustment", "Occupancy Adj", False))
        else:
            pnl_rows.append(("Vacancy Loss", "Vacancy Loss", False))
        
        pnl_rows.extend([
            ("Gross Operating Income", "Gross Operating Income", True),
            ("Property Tax", "Property Tax", False),
            ("Condo Fees", "Condo Fees", False),
            ("PNO Insurance", "PNO Insurance", False),
            ("Maintenance", "Maintenance", False),
            ("Management Fees", "Management Fees", False),
        ])
        
        if self.lease_type == "airbnb":
            pnl_rows.append(("Airbnb Costs", "Airbnb Costs", False))
        
        pnl_rows.extend([
            ("Total Operating Expenses", "Total OpEx", True),
            ("Net Operating Income", "NOI", True),
            ("Loan Interest", "Loan Interest", False),
            ("Loan Insurance", "Loan Insurance", False),
            ("Depreciation - Property", "Depr Property", False),
            ("Depreciation - Furnishing", "Depr Furnishing", False),
            ("Depreciation - Renovation", "Depr Renovation", False),
            ("Total Depreciation", "Total Depreciation", True),
            ("Taxable Income", "Taxable Income", False),
            ("Income Tax", "Income Tax", False),
            ("Social Contributions", "Social Contributions", False),
            ("Total Taxes", "Total Taxes", True),
            ("Net Income", "Net Income", True),
        ])
        
        # Header row
        ws.cell(row=1, column=1, value="P&L Summary (€)")
        ws.cell(row=1, column=1).font = self.HEADER_FONT
        ws.cell(row=1, column=1).fill = self.HEADER_FILL
        
        for year in range(1, years + 1):
            cell = ws.cell(row=1, column=year + 1, value=f"Year {year}")
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows with SUMIF formulas
        current_row = 2
        for label, monthly_col_name, is_total in pnl_rows:
            ws.cell(row=current_row, column=1, value=label)
            
            # Get column letter from map
            col_num = self.pnl_col_map.get(monthly_col_name, None)
            if col_num:
                col_letter = get_column_letter(col_num)
                
                for year in range(1, years + 1):
                    # SUMIF: sum monthly column where Year = this year
                    formula = f"=SUMIF(PnL_Monthly!$B:$B,{year},PnL_Monthly!${col_letter}:${col_letter})"
                    cell = ws.cell(row=current_row, column=year + 1, value=formula)
                    cell.number_format = self.NUM_FORMAT
                    
                    if is_total:
                        cell.font = self.BOLD_TOTAL
                        cell.fill = self.SUBTOTAL_FILL
            
            if is_total:
                ws.cell(row=current_row, column=1).font = self.BOLD_TOTAL
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 28
        for col in range(2, years + 2):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_yearly_cf_sheet(self):
        """Create yearly Cash Flow aggregating monthly data."""
        ws = self.wb.create_sheet("CF_Yearly")
        
        years = self.params.holding_period_years
        
        cf_cols = [
            ("Net Income", "C", False),
            ("Add: Depreciation", "D", False),
            ("Cash Flow from Operations", "E", True),
            ("Acquisition Costs", "F", False),
            ("Cash Flow from Investing", "G", True),
            ("Loan Proceeds", "H", False),
            ("Equity Injected", "I", False),
            ("Principal Repayment", "J", False),
            ("Cash Flow from Financing", "K", True),
            ("Net Change in Cash", "L", True),
        ]
        
        # Headers
        ws.cell(row=1, column=1, value="Cash Flow Summary (€)")
        ws.cell(row=1, column=1).font = self.HEADER_FONT
        ws.cell(row=1, column=1).fill = self.HEADER_FILL
        
        for year in range(1, years + 1):
            cell = ws.cell(row=1, column=year + 1, value=f"Year {year}")
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        
        current_row = 2
        for label, monthly_col, is_total in cf_cols:
            ws.cell(row=current_row, column=1, value=label)
            
            for year in range(1, years + 1):
                formula = f"=SUMIF(CF_Monthly!$B:$B,{year},CF_Monthly!${monthly_col}:${monthly_col})"
                cell = ws.cell(row=current_row, column=year + 1, value=formula)
                cell.number_format = self.NUM_FORMAT
                
                if is_total:
                    cell.font = self.BOLD_TOTAL
                    cell.fill = self.SUBTOTAL_FILL
            
            current_row += 1
        
        # Beginning/Ending Cash (not summed - take first/last of year)
        ws.cell(row=current_row, column=1, value="Beginning Cash")
        for year in range(1, years + 1):
            start_month = (year - 1) * 12 + 1
            formula = f"=CF_Monthly!M{start_month + 1}"  # +1 for header row
            ws.cell(row=current_row, column=year + 1, value=formula).number_format = self.NUM_FORMAT
        current_row += 1
        
        ws.cell(row=current_row, column=1, value="Ending Cash")
        ws.cell(row=current_row, column=1).font = self.BOLD_TOTAL
        for year in range(1, years + 1):
            end_month = year * 12
            formula = f"=CF_Monthly!N{end_month + 1}"
            cell = ws.cell(row=current_row, column=year + 1, value=formula)
            cell.number_format = self.NUM_FORMAT
            cell.font = self.BOLD_TOTAL
            cell.fill = self.TOTAL_FILL
        
        ws.column_dimensions['A'].width = 28
        for col in range(2, years + 2):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_yearly_bs_sheet(self):
        """Create yearly Balance Sheet (year-end snapshots)."""
        ws = self.wb.create_sheet("BS_Yearly")
        
        years = self.params.holding_period_years
        
        bs_cols = [
            ("ASSETS", None, True),
            ("Property Net Value", "E", False),
            ("Renovation Net Value", "H", False),
            ("Furnishing Net Value", "K", False),
            ("Total Fixed Assets", "L", True),
            ("Cash", "M", False),
            ("Total Assets", "N", True),
            ("", None, False),
            ("LIABILITIES", None, True),
            ("Loan Balance", "O", False),
            ("Total Liabilities", "P", True),
            ("", None, False),
            ("EQUITY", None, True),
            ("Initial Equity", "Q", False),
            ("Retained Earnings", "R", False),
            ("Total Equity", "S", True),
            ("", None, False),
            ("Total Liabilities & Equity", "T", True),
            ("Balance Check", "U", True),
        ]
        
        # Headers
        ws.cell(row=1, column=1, value="Balance Sheet (€)")
        ws.cell(row=1, column=1).font = self.HEADER_FONT
        ws.cell(row=1, column=1).fill = self.HEADER_FILL
        
        ws.cell(row=1, column=2, value="Initial")
        ws.cell(row=1, column=2).font = self.HEADER_FONT
        ws.cell(row=1, column=2).fill = self.HEADER_FILL
        
        for year in range(1, years + 1):
            cell = ws.cell(row=1, column=year + 2, value=f"Year {year}")
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
        
        current_row = 2
        for label, monthly_col, is_section in bs_cols:
            if label == "":
                current_row += 1
                continue
            
            ws.cell(row=current_row, column=1, value=label)
            
            if monthly_col:
                # Initial (Month 0 = row 2 in BS_Monthly)
                formula = f"=BS_Monthly!{monthly_col}2"
                ws.cell(row=current_row, column=2, value=formula).number_format = self.NUM_FORMAT
                
                # Year-end values
                for year in range(1, years + 1):
                    month_row = year * 12 + 2  # Month 12, 24, etc. + 2 for header and month 0
                    formula = f"=BS_Monthly!{monthly_col}{month_row}"
                    cell = ws.cell(row=current_row, column=year + 2, value=formula)
                    cell.number_format = self.NUM_FORMAT
            
            if is_section or "Total" in label:
                for col in range(1, years + 3):
                    ws.cell(row=current_row, column=col).font = self.BOLD_TOTAL
                    if "Total" in label:
                        ws.cell(row=current_row, column=col).fill = self.SUBTOTAL_FILL
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 28
        for col in range(2, years + 3):
            ws.column_dimensions[get_column_letter(col)].width = 14
    
    def _create_metrics_sheet(self):
        """Create Investment Metrics with formulas where possible."""
        ws = self.wb.create_sheet("Metrics")
        
        years = self.params.holding_period_years
        
        metrics = [
            ("INVESTMENT METRICS", None, None),
            ("IRR (requires manual XIRR)", self.metrics.get('irr', 0), self.PCT_FORMAT),
            ("NPV (requires manual NPV)", self.metrics.get('npv', 0), self.NUM_FORMAT),
            ("Cash-on-Cash Year 1", f"=CF_Yearly!L2/InitialEquity", self.PCT_FORMAT),
            ("", None, None),
            ("EXIT SCENARIO (Year {})".format(years), None, None),
            ("Exit Property Value", f"=PropertyPrice*(1+PropertyGrowth)^HoldingYears", self.NUM_FORMAT),
            ("Selling Costs", "=-B7*SellingFeesPct", self.NUM_FORMAT),
            ("Net Selling Price", "=B7+B8", self.NUM_FORMAT),
            ("Remaining Loan Balance", f"=BS_Monthly!O{years*12+2}", self.NUM_FORMAT),
            ("Gross Proceeds", "=B9-B10", self.NUM_FORMAT),
            ("Capital Gains Tax (simplified)", self.metrics.get('capital_gains_tax', 0), self.NUM_FORMAT),
            ("Net Exit Proceeds", "=B11-B12", self.NUM_FORMAT),
            ("", None, None),
            ("EQUITY MULTIPLE", None, None),
            ("Total Cash Flows", f"=SUM(CF_Yearly!L2:L{years+1})", self.NUM_FORMAT),
            ("Plus Exit Proceeds", "=B13", self.NUM_FORMAT),
            ("Total Return", "=B16+B17", self.NUM_FORMAT),
            ("Equity Multiple", "=B18/InitialEquity", "0.00x"),
        ]
        
        row = 1
        for label, value, fmt in metrics:
            if value is None and label and label != "":
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=1).font = self.HEADER_FONT
                ws.cell(row=row, column=1).fill = self.HEADER_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            elif label == "":
                pass
            else:
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=value)
                if fmt:
                    cell.number_format = fmt
                if "Multiple" in label or "Proceeds" in label:
                    cell.font = self.BOLD_TOTAL
                    cell.fill = self.TOTAL_FILL
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 18
