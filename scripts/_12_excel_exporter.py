# In file: scripts/_12_excel_exporter.py

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Optional
from io import BytesIO


class ExcelExporter:
    """
    Exports financial model to Excel with formulas, formatting, and linked sheets.
    Follows accounting conventions: brackets for negatives, sums only for totals.
    """
    
    # Style constants
    BLUE_INPUT = Font(color="0000FF", bold=False)
    BLACK_FORMULA = Font(color="000000", bold=False)
    BOLD = Font(bold=True)
    BOLD_TOTAL = Font(bold=True, color="000000")
    
    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
    SUBTOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
    TOTAL_FILL = PatternFill("solid", fgColor="BDD7EE")
    
    THIN_BORDER = Border(bottom=Side(style='thin', color='000000'))
    THICK_BORDER = Border(top=Side(style='medium', color='000000'), bottom=Side(style='double', color='000000'))
    
    # Number formats (no currency symbol, brackets for negatives)
    NUM_FORMAT = '#,##0;(#,##0);"-"'
    NUM_FORMAT_DEC = '#,##0.00;(#,##0.00);"-"'
    PCT_FORMAT = '0.0%;(0.0%);"-"'
    
    def __init__(self, params, pnl_df: pd.DataFrame, bs_df: pd.DataFrame, 
                 cf_df: pd.DataFrame, loan_schedule: pd.DataFrame,
                 investment_metrics: Dict):
        self.params = params
        self.pnl_df = pnl_df
        self.bs_df = bs_df
        self.cf_df = cf_df
        self.loan_schedule = loan_schedule
        self.metrics = investment_metrics or {}
        self.wb = Workbook()
        
    def export(self) -> BytesIO:
        """Generate Excel workbook and return as BytesIO for download."""
        self.wb.remove(self.wb.active)
        
        self._create_assumptions_sheet()
        self._create_pnl_sheet()
        self._create_bs_sheet()
        self._create_cf_sheet()
        self._create_loan_sheet()
        self._create_metrics_sheet()
        
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    def _create_assumptions_sheet(self):
        """Create Assumptions tab with all input parameters."""
        ws = self.wb.create_sheet("Assumptions")
        
        assumptions = [
            ("PROPERTY & ACQUISITION", None),
            ("Property Price (FAI)", self.params.property_price),
            ("Property Size (sqm)", self.params.property_size_sqm),
            ("Agency Fees %", self.params.agency_fees_percentage),
            ("Notary Fees %", self.params.notary_fees_percentage_estimate),
            ("Initial Renovation", self.params.initial_renovation_costs),
            ("Furnishing Costs", self.params.furnishing_costs),
            ("", None),
            ("FINANCING", None),
            ("Loan Percentage", self.params.loan_percentage),
            ("Loan Interest Rate", self.params.loan_interest_rate),
            ("Loan Duration (Years)", self.params.loan_duration_years),
            ("Loan Insurance Rate", self.params.loan_insurance_rate),
            ("", None),
            ("OPERATING EXPENSES", None),
            ("Property Tax (Yearly)", self.params.property_tax_yearly),
            ("Condo Fees (Monthly)", self.params.condo_fees_monthly),
            ("PNO Insurance (Yearly)", self.params.pno_insurance_yearly),
            ("Maintenance % of Rent", self.params.maintenance_percentage_rent),
            ("Expenses Growth Rate", self.params.expenses_growth_rate),
            ("", None),
            ("FISCAL PARAMETERS", None),
            ("Fiscal Regime", self.params.fiscal_regime),
            ("Income Tax Bracket (TMI)", self.params.personal_income_tax_bracket),
            ("Social Contributions Rate", self.params.social_contributions_rate),
            ("", None),
            ("EXIT PARAMETERS", None),
            ("Holding Period (Years)", self.params.holding_period_years),
            ("Property Growth Rate", self.params.property_value_growth_rate),
            ("Selling Fees %", self.params.exit_selling_fees_percentage),
            ("", None),
            ("INVESTMENT ANALYSIS", None),
            ("Risk-Free Rate", getattr(self.params, 'risk_free_rate', 0.035)),
            ("Discount Rate", getattr(self.params, 'discount_rate', 0.05)),
        ]
        
        row = 1
        for label, value in assumptions:
            if value is None and label and label != "":
                # Section header
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=1).font = self.HEADER_FONT
                ws.cell(row=row, column=1).fill = self.HEADER_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            elif label == "":
                pass
            else:
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=value)
                cell.font = self.BLUE_INPUT
                if isinstance(value, float):
                    if "%" in label or "Rate" in label:
                        cell.number_format = self.PCT_FORMAT
                    else:
                        cell.number_format = self.NUM_FORMAT
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        
    def _create_pnl_sheet(self):
        """Create P&L tab with yearly data and formulas."""
        ws = self.wb.create_sheet("P&L")
        
        pnl_yearly = self.pnl_df.groupby("Year").sum()
        years = list(pnl_yearly.index)
        
        # Row definitions: (label, df_column, is_expense, formula_type)
        pnl_rows = [
            ("Gross Potential Rent", "Gross Potential Rent", False, None),
            ("Vacancy Loss", "Vacancy Loss", True, None),
            ("Gross Operating Income", "Gross Operating Income", False, "GOI"),
            ("", None, False, None),
            ("Property Tax", "Property Tax", True, None),
            ("Condo Fees", "Condo Fees", True, None),
            ("PNO Insurance", "PNO Insurance", True, None),
            ("Maintenance", "Maintenance", True, None),
            ("Management Fees", "Management Fees", True, None),
            ("Total Operating Expenses", "Total Operating Expenses", True, "OPEX"),
            ("Net Operating Income", "Net Operating Income", False, "NOI"),
            ("", None, False, None),
            ("Loan Interest", "Loan Interest", True, None),
            ("Loan Insurance", "Loan Insurance", True, None),
            ("Depreciation/Amortization", "Depreciation/Amortization", True, None),
            ("Taxable Income", "Taxable Income", False, "TAXABLE"),
            ("", None, False, None),
            ("Income Tax", "Income Tax", True, None),
            ("Social Contributions", "Social Contributions", True, None),
            ("Total Taxes", "Total Taxes", True, "TAX_TOTAL"),
            ("Net Income", "Net Income", False, "NET"),
        ]
        
        # Header row
        ws.cell(row=1, column=1, value="P&L Statement (€)")
        ws.cell(row=1, column=1).font = self.HEADER_FONT
        ws.cell(row=1, column=1).fill = self.HEADER_FILL
        
        for col_idx, year in enumerate(years, start=2):
            cell = ws.cell(row=1, column=col_idx, value=f"Year {year}")
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        row_map = {}
        current_row = 2
        
        for label, df_col, is_expense, formula_type in pnl_rows:
            if label == "":
                current_row += 1
                continue
            
            ws.cell(row=current_row, column=1, value=label)
            row_map[label] = current_row
            
            for col_idx, year in enumerate(years, start=2):
                col_letter = get_column_letter(col_idx)
                
                if formula_type == "GOI":
                    formula = f"={col_letter}{row_map['Gross Potential Rent']}+{col_letter}{row_map['Vacancy Loss']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "OPEX":
                    formula = f"=SUM({col_letter}{row_map['Property Tax']}:{col_letter}{row_map['Management Fees']})"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "NOI":
                    formula = f"={col_letter}{row_map['Gross Operating Income']}+{col_letter}{row_map['Total Operating Expenses']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "TAXABLE":
                    formula = f"={col_letter}{row_map['Net Operating Income']}+{col_letter}{row_map['Loan Interest']}+{col_letter}{row_map['Loan Insurance']}+{col_letter}{row_map['Depreciation/Amortization']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "TAX_TOTAL":
                    formula = f"=SUM({col_letter}{row_map['Income Tax']}:{col_letter}{row_map['Social Contributions']})"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "NET":
                    formula = f"={col_letter}{row_map['Taxable Income']}+{col_letter}{row_map['Total Taxes']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                else:
                    if df_col and df_col in pnl_yearly.columns:
                        value = pnl_yearly.loc[year, df_col]
                        if is_expense:
                            value = -abs(value) if value != 0 else 0
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                    else:
                        cell = ws.cell(row=current_row, column=col_idx, value=0)
                
                cell.number_format = self.NUM_FORMAT
                if formula_type:
                    cell.font = self.BLACK_FORMULA
            
            if "Total" in label or label in ["Gross Operating Income", "Net Operating Income", "Taxable Income", "Net Income"]:
                for c in range(1, len(years) + 2):
                    ws.cell(row=current_row, column=c).font = self.BOLD_TOTAL
                    if "Total" in label:
                        ws.cell(row=current_row, column=c).fill = self.SUBTOTAL_FILL
                    if label == "Net Income":
                        ws.cell(row=current_row, column=c).fill = self.TOTAL_FILL
                        ws.cell(row=current_row, column=c).border = self.THICK_BORDER
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 30
        for col_idx in range(2, len(years) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    def _create_bs_sheet(self):
        """Create Balance Sheet tab with formulas."""
        ws = self.wb.create_sheet("Balance Sheet")
        
        key_months = [0] + [12 * y for y in range(1, self.params.holding_period_years + 1)]
        bs_yearly = self.bs_df.loc[self.bs_df.index.intersection(key_months)]
        
        bs_rows = [
            ("ASSETS", None, True, None),
            ("Property Net Value", "Property Net Value", False, None),
            ("Renovation Net Value", "Renovation Net Value", False, None),
            ("Furnishing Net Value", "Furnishing Net Value", False, None),
            ("Total Fixed Assets", "Total Fixed Assets", True, "SUM_ASSETS"),
            ("Cash", "Cash", False, None),
            ("Total Assets", "Total Assets", True, "TOTAL_ASSETS"),
            ("", None, False, None),
            ("LIABILITIES", None, True, None),
            ("Loan Balance", "Loan Balance", False, None),
            ("Total Liabilities", "Total Liabilities", True, "TOTAL_LIAB"),
            ("", None, False, None),
            ("EQUITY", None, True, None),
            ("Initial Equity", "Initial Equity", False, None),
            ("Retained Earnings", "Retained Earnings", False, None),
            ("Total Equity", "Total Equity", True, "TOTAL_EQUITY"),
            ("", None, False, None),
            ("Total Liabilities & Equity", "Total Liabilities and Equity", True, "TOTAL_LE"),
            ("Balance Check", "Balance Check", True, "CHECK"),
        ]
        
        ws.cell(row=1, column=1, value="Balance Sheet (€)")
        ws.cell(row=1, column=1).font = self.HEADER_FONT
        ws.cell(row=1, column=1).fill = self.HEADER_FILL
        
        col_idx = 2
        for month in bs_yearly.index:
            label = "Initial" if month == 0 else f"Year {month // 12}"
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            col_idx += 1
        
        row_map = {}
        current_row = 2
        
        for label, df_col, is_total, formula_type in bs_rows:
            if label == "":
                current_row += 1
                continue
            
            ws.cell(row=current_row, column=1, value=label)
            row_map[label] = current_row
            
            col_idx = 2
            for month in bs_yearly.index:
                col_letter = get_column_letter(col_idx)
                
                if formula_type == "SUM_ASSETS":
                    formula = f"=SUM({col_letter}{row_map['Property Net Value']}:{col_letter}{row_map['Furnishing Net Value']})"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "TOTAL_ASSETS":
                    formula = f"={col_letter}{row_map['Total Fixed Assets']}+{col_letter}{row_map['Cash']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "TOTAL_LIAB":
                    formula = f"={col_letter}{row_map['Loan Balance']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "TOTAL_EQUITY":
                    formula = f"={col_letter}{row_map['Initial Equity']}+{col_letter}{row_map['Retained Earnings']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "TOTAL_LE":
                    formula = f"={col_letter}{row_map['Total Liabilities']}+{col_letter}{row_map['Total Equity']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "CHECK":
                    formula = f"={col_letter}{row_map['Total Assets']}-{col_letter}{row_map['Total Liabilities & Equity']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif df_col and df_col in bs_yearly.columns:
                    value = bs_yearly.loc[month, df_col]
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                else:
                    cell = ws.cell(row=current_row, column=col_idx, value=0)
                
                cell.number_format = self.NUM_FORMAT
                col_idx += 1
            
            if is_total or label in ["ASSETS", "LIABILITIES", "EQUITY"]:
                for c in range(1, len(bs_yearly) + 2):
                    ws.cell(row=current_row, column=c).font = self.BOLD_TOTAL
                    if "Total" in label:
                        ws.cell(row=current_row, column=c).fill = self.SUBTOTAL_FILL
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 30
        for col_idx in range(2, len(bs_yearly) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    def _create_cf_sheet(self):
        """Create Cash Flow tab with formulas."""
        ws = self.wb.create_sheet("Cash Flow")
        
        cf_yearly = self.cf_df.groupby("Year").sum()
        
        for year in cf_yearly.index:
            year_data = self.cf_df[self.cf_df["Year"] == year]
            cf_yearly.loc[year, "Beginning Cash Balance"] = year_data["Beginning Cash Balance"].iloc[0]
            cf_yearly.loc[year, "Ending Cash Balance"] = year_data["Ending Cash Balance"].iloc[-1]
        
        years = list(cf_yearly.index)
        
        cf_rows = [
            ("OPERATING ACTIVITIES", None, True, None),
            ("Net Income", "Net Income", False, None),
            ("Depreciation/Amortization", "Depreciation/Amortization", False, None),
            ("Cash Flow from Operations", "Cash Flow from Operations (CFO)", True, "CFO"),
            ("", None, False, None),
            ("INVESTING ACTIVITIES", None, True, None),
            ("Acquisition Costs", "Acquisition Costs Outflow", False, None),
            ("Cash Flow from Investing", "Cash Flow from Investing (CFI)", True, "CFI"),
            ("", None, False, None),
            ("FINANCING ACTIVITIES", None, True, None),
            ("Loan Proceeds", "Loan Proceeds", False, None),
            ("Equity Injected", "Equity Injected", False, None),
            ("Loan Principal Repayment", "Loan Principal Repayment", False, None),
            ("Cash Flow from Financing", "Cash Flow from Financing (CFF)", True, "CFF"),
            ("", None, False, None),
            ("Net Change in Cash", "Net Change in Cash", True, "NET_CHANGE"),
            ("Beginning Cash", "Beginning Cash Balance", False, None),
            ("Ending Cash", "Ending Cash Balance", True, "END_CASH"),
        ]
        
        ws.cell(row=1, column=1, value="Cash Flow Statement (€)")
        ws.cell(row=1, column=1).font = self.HEADER_FONT
        ws.cell(row=1, column=1).fill = self.HEADER_FILL
        
        for col_idx, year in enumerate(years, start=2):
            cell = ws.cell(row=1, column=col_idx, value=f"Year {year}")
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        row_map = {}
        current_row = 2
        
        for label, df_col, is_total, formula_type in cf_rows:
            if label == "":
                current_row += 1
                continue
            
            ws.cell(row=current_row, column=1, value=label)
            row_map[label] = current_row
            
            for col_idx, year in enumerate(years, start=2):
                col_letter = get_column_letter(col_idx)
                
                if formula_type == "CFO":
                    formula = f"={col_letter}{row_map['Net Income']}+{col_letter}{row_map['Depreciation/Amortization']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "CFI":
                    formula = f"={col_letter}{row_map['Acquisition Costs']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "CFF":
                    formula = f"={col_letter}{row_map['Loan Proceeds']}+{col_letter}{row_map['Equity Injected']}+{col_letter}{row_map['Loan Principal Repayment']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "NET_CHANGE":
                    formula = f"={col_letter}{row_map['Cash Flow from Operations']}+{col_letter}{row_map['Cash Flow from Investing']}+{col_letter}{row_map['Cash Flow from Financing']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                elif formula_type == "END_CASH":
                    formula = f"={col_letter}{row_map['Beginning Cash']}+{col_letter}{row_map['Net Change in Cash']}"
                    cell = ws.cell(row=current_row, column=col_idx, value=formula)
                else:
                    if df_col and df_col in cf_yearly.columns:
                        value = cf_yearly.loc[year, df_col]
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                    else:
                        cell = ws.cell(row=current_row, column=col_idx, value=0)
                
                cell.number_format = self.NUM_FORMAT
            
            if is_total or "ACTIVITIES" in label:
                for c in range(1, len(years) + 2):
                    ws.cell(row=current_row, column=c).font = self.BOLD_TOTAL
                    if is_total and "ACTIVITIES" not in label:
                        ws.cell(row=current_row, column=c).fill = self.SUBTOTAL_FILL
            
            current_row += 1
        
        ws.column_dimensions['A'].width = 30
        for col_idx in range(2, len(years) + 2):
            ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    def _create_loan_sheet(self):
        """Create Loan Amortization tab."""
        ws = self.wb.create_sheet("Loan Schedule")
        
        if self.loan_schedule is None or len(self.loan_schedule) == 0:
            ws.cell(row=1, column=1, value="No loan in this scenario (100% equity)")
            return
        
        loan_copy = self.loan_schedule.copy()
        loan_copy['Year'] = ((loan_copy.index - 1) // 12) + 1
        
        yearly = loan_copy.groupby('Year').agg({
            'Beginning Balance': 'first',
            'Monthly Payment': 'sum',
            'Interest Payment': 'sum',
            'Principal Payment': 'sum',
            'Ending Balance': 'last'
        })
        
        headers = ["Year", "Beginning Balance", "Total Payments", "Interest", "Principal", "Ending Balance"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, (year, data) in enumerate(yearly.iterrows(), start=2):
            ws.cell(row=row_idx, column=1, value=year)
            ws.cell(row=row_idx, column=2, value=data['Beginning Balance']).number_format = self.NUM_FORMAT
            ws.cell(row=row_idx, column=3, value=data['Monthly Payment']).number_format = self.NUM_FORMAT
            ws.cell(row=row_idx, column=4, value=data['Interest Payment']).number_format = self.NUM_FORMAT
            ws.cell(row=row_idx, column=5, value=data['Principal Payment']).number_format = self.NUM_FORMAT
            ws.cell(row=row_idx, column=6, value=data['Ending Balance']).number_format = self.NUM_FORMAT
        
        total_row = len(yearly) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = self.BOLD_TOTAL
        ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{total_row-1})").number_format = self.NUM_FORMAT
        ws.cell(row=total_row, column=4, value=f"=SUM(D2:D{total_row-1})").number_format = self.NUM_FORMAT
        ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row-1})").number_format = self.NUM_FORMAT
        
        for col_idx in range(1, 7):
            ws.cell(row=total_row, column=col_idx).fill = self.TOTAL_FILL
            ws.cell(row=total_row, column=col_idx).font = self.BOLD_TOTAL
        
        for col_idx in range(1, 7):
            ws.column_dimensions[get_column_letter(col_idx)].width = 18
    
    def _create_metrics_sheet(self):
        """Create Investment Metrics tab."""
        ws = self.wb.create_sheet("Investment Metrics")
        
        metrics_data = [
            ("INVESTMENT PERFORMANCE", None, None),
            ("IRR (Annual)", self.metrics.get('irr', 0), self.PCT_FORMAT),
            ("NPV", self.metrics.get('npv', 0), self.NUM_FORMAT),
            ("Cash-on-Cash (Year 1)", self.metrics.get('cash_on_cash', 0), self.PCT_FORMAT),
            ("Equity Multiple", self.metrics.get('equity_multiple', 0), '0.00x'),
            ("", None, None),
            ("EXIT SCENARIO", None, None),
            ("Exit Property Value", self.metrics.get('exit_property_value', 0), self.NUM_FORMAT),
            ("Selling Costs", self.metrics.get('selling_costs', 0), self.NUM_FORMAT),
            ("Remaining Loan Balance", self.metrics.get('remaining_loan_balance', 0), self.NUM_FORMAT),
            ("Capital Gain", self.metrics.get('capital_gain', 0), self.NUM_FORMAT),
            ("Capital Gains Tax", self.metrics.get('capital_gains_tax', 0), self.NUM_FORMAT),
            ("Net Exit Proceeds", self.metrics.get('net_exit_proceeds', 0), self.NUM_FORMAT),
        ]
        
        row = 1
        for label, value, fmt in metrics_data:
            if value is None and label and label != "":
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=1).font = self.HEADER_FONT
                ws.cell(row=row, column=1).fill = self.HEADER_FILL
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            elif label == "":
                pass
            else:
                ws.cell(row=row, column=1, value=label)
                cell = ws.cell(row=row, column=2, value=value)
                if fmt:
                    cell.number_format = fmt
                if label in ["Net Exit Proceeds", "Equity Multiple"]:
                    ws.cell(row=row, column=1).font = self.BOLD_TOTAL
                    cell.font = self.BOLD_TOTAL
                    cell.fill = self.TOTAL_FILL
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18
